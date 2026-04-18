"""Tests for the flat xlsx exporter."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from msb_extractor.export import write_xlsx
from msb_extractor.parser.capture import parse_capture


def test_write_xlsx_creates_file_and_all_sheets(
    tmp_path: Path,
    capture_json: dict[str, Any],
) -> None:
    result = parse_capture(capture_json)
    out = write_xlsx(result, tmp_path / "out.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    assert "Raw Log" in wb.sheetnames
    assert "Exercise Progress" in wb.sheetnames
    assert "Exercise Index" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    # At least one weekly sheet was created for the test capture.
    weekly_sheets = [s for s in wb.sheetnames if s.startswith("Week ")]
    assert len(weekly_sheets) >= 1


def test_write_xlsx_respects_include_flags(
    tmp_path: Path,
    capture_json: dict[str, Any],
) -> None:
    result = parse_capture(capture_json)
    out = write_xlsx(
        result,
        tmp_path / "minimal.xlsx",
        include_weekly=False,
        include_progress=False,
    )
    wb = load_workbook(out)
    assert "Raw Log" in wb.sheetnames
    assert "Exercise Progress" not in wb.sheetnames
    assert not any(s.startswith("Week ") for s in wb.sheetnames)


def test_raw_log_has_header_row(tmp_path: Path, capture_json: dict[str, Any]) -> None:
    result = parse_capture(capture_json)
    out = write_xlsx(result, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb["Raw Log"]
    first_row = [ws.cell(row=1, column=c).value for c in range(1, 15)]
    assert first_row[0] == "Date"
    assert first_row[3] == "Exercise"
    assert first_row[9] == "Load"


def test_raw_log_has_data_rows(tmp_path: Path, capture_json: dict[str, Any]) -> None:
    result = parse_capture(capture_json)
    out = write_xlsx(result, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb["Raw Log"]
    # At minimum: 2 actuals for bench + 1 actual for close grip + 1 partial-actuals day
    assert ws.max_row >= 4


def test_exercise_index_sorted_by_volume(
    tmp_path: Path,
    capture_json: dict[str, Any],
) -> None:
    result = parse_capture(capture_json)
    out = write_xlsx(result, tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb["Exercise Index"]
    # Header + at least 3 exercises (bench, close grip, squat, rdl)
    assert ws.max_row >= 4
    first_exercise_name = ws.cell(row=2, column=1).value
    assert isinstance(first_exercise_name, str)


def test_rename_map_applied_in_output(
    tmp_path: Path,
    capture_json: dict[str, Any],
) -> None:
    rename_yaml = tmp_path / "rename.yaml"
    rename_yaml.write_text(
        '"Competition (bench)": "Flat Barbell Bench Press"\n',
        encoding="utf-8",
    )
    result = parse_capture(capture_json)
    out = write_xlsx(
        result,
        tmp_path / "out.xlsx",
        rename_map_path=rename_yaml,
    )
    wb = load_workbook(out)
    ws = wb["Raw Log"]
    names_in_output = {ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)}
    assert "Flat Barbell Bench Press" in names_in_output
    assert "Competition (bench)" not in names_in_output

"""Weekly-view sheet writer: one sheet per ISO week, wide format.

The weekly view is the presentation layer modelled on lifter-facing training
logs like Jeff Nippard's Powerbuilding and Peace Era PPL spreadsheets:
exercises are rows, sets are columns, one block per training day, one sheet
per week.

This is the sheet most users will spend their time in.
"""

from __future__ import annotations

from datetime import date as date_type

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from msb_extractor.export._styles import (
    DAY_BANNER_FILL,
    DAY_BANNER_FONT,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    STATUS_COLORS,
    TITLE_FONT,
    WRAP_ALIGN,
)
from msb_extractor.models import Exercise, ParseResult, SetStatus, TrainingDay
from msb_extractor.normalize.exercise import apply_rename
from msb_extractor.normalize.program import group_by_week
from msb_extractor.normalize.units import Unit, format_load

_DEFAULT_MAX_SETS = 5


def write_weekly_view(
    wb: Workbook,
    result: ParseResult,
    unit: Unit,
    rename_map: dict[str, str],
    max_sets: int = _DEFAULT_MAX_SETS,
) -> None:
    """Create one sheet per training week present in ``result``."""
    buckets = group_by_week(result.days)
    for week_start, days in buckets.items():
        sheet_name = _sheet_name_for_week(week_start)
        ws = wb.create_sheet(sheet_name)
        _write_week_sheet(ws, week_start, days, unit, rename_map, max_sets)


def _sheet_name_for_week(week_start: date_type) -> str:
    # Sheet names max out at 31 chars in Excel; this is comfortably within.
    return f"Week {week_start.isoformat()}"


def _write_week_sheet(
    ws: Worksheet,
    week_start: date_type,
    days: list[TrainingDay],
    unit: Unit,
    rename_map: dict[str, str],
    max_sets: int,
) -> None:
    total_cols = 4 + max_sets + 2  # exercise, sets, reps, rpe, setsN..., lsrpe, notes
    last_col_letter = get_column_letter(total_cols)

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"Week of {week_start.isoformat()}")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Column widths
    widths = [32, 7, 9, 7] + [14] * max_sets + [8, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for day in days:
        row = _write_day_section(ws, row, day, unit, rename_map, max_sets, last_col_letter)
        row += 1  # blank spacer row

    ws.freeze_panes = "A3"


def _write_day_section(
    ws: Worksheet,
    start_row: int,
    day: TrainingDay,
    unit: Unit,
    rename_map: dict[str, str],
    max_sets: int,
    last_col_letter: str,
) -> int:
    row = start_row
    total_cols = int(last_col_letter_to_number(last_col_letter))

    # Day banner
    banner_text = f"{day.date.strftime('%A')}  {day.date.isoformat()}"
    banner_cell = ws.cell(row=row, column=1, value=banner_text)
    banner_cell.font = DAY_BANNER_FONT
    banner_cell.fill = DAY_BANNER_FILL
    banner_cell.alignment = HEADER_ALIGN
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    row += 1

    # Column headers
    headers = ["Exercise", "Sets", "Reps", "RPE"]
    headers.extend(f"Set {i}" for i in range(1, max_sets + 1))
    headers.extend(["LSRPE", "Notes"])
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    row += 1

    # Exercise rows
    for ex in day.exercises:
        display_name = apply_rename(ex.name, rename_map)
        _write_exercise_row(ws, row, ex, display_name, unit, max_sets)
        row += 1

    return row


def last_col_letter_to_number(letter: str) -> int:
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch.upper()) - ord("A") + 1)
    return result


def _write_exercise_row(
    ws: Worksheet,
    row: int,
    ex: Exercise,
    display_name: str,
    unit: Unit,
    max_sets: int,
) -> None:
    label = f"{ex.order}. {display_name}" if ex.order else display_name
    ws.cell(row=row, column=1, value=label)

    first_pres = ex.prescribed[0] if ex.prescribed else None
    prescribed_sets = first_pres.sets if first_pres and first_pres.sets else None
    if prescribed_sets is None and ex.actuals:
        prescribed_sets = len(ex.actuals)

    prescribed_reps: int | str | None = None
    if first_pres:
        prescribed_reps = first_pres.reps if first_pres.reps else first_pres.reps_text

    ws.cell(row=row, column=2, value=prescribed_sets)
    ws.cell(row=row, column=3, value=prescribed_reps)
    rpe_cell = ws.cell(row=row, column=4, value=first_pres.rpe if first_pres else None)
    if first_pres and first_pres.rpe is not None:
        rpe_cell.number_format = "0.0"

    # Per-set loads (prefer actuals, fall back to prescribed)
    entries = ex.actuals if ex.actuals else ex.prescribed
    target_reps = first_pres.reps if first_pres else None

    for i, entry in enumerate(entries[:max_sets]):
        col = 5 + i
        load_kg = getattr(entry, "load_kg", None)
        load_display = getattr(entry, "load_display", None)
        actual_reps = getattr(entry, "reps", None)

        cell_value: str | None = None
        if load_kg is not None:
            cell_value = format_load(load_kg, unit, decimals=1)
        elif load_display:
            cell_value = load_display

        # Annotate when reps fell short of the prescription (example convention).
        if (
            cell_value
            and target_reps
            and isinstance(actual_reps, int)
            and actual_reps < target_reps
        ):
            cell_value = f"{cell_value} x {actual_reps}"

        if cell_value:
            cell = ws.cell(row=row, column=col, value=cell_value)
            status = getattr(entry, "status", SetStatus.UNKNOWN)
            if status == SetStatus.MISSED:
                cell.font = Font(color=STATUS_COLORS[SetStatus.MISSED], bold=True)

    # LSRPE: actual RPE on the last logged set
    if ex.actuals:
        last = ex.actuals[-1]
        if last.rpe is not None:
            lsrpe_cell = ws.cell(row=row, column=5 + max_sets, value=last.rpe)
            lsrpe_cell.number_format = "0.0"

    # Notes: concatenate every non-empty per-set comment
    notes = [f"*{a.comment}" for a in ex.actuals if a.comment]
    if notes:
        notes_cell = ws.cell(row=row, column=6 + max_sets, value=" | ".join(notes))
        notes_cell.alignment = WRAP_ALIGN

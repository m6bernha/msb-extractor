"""Flat sheet writers: Raw Log, Exercise Index, Summary."""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from typing import Any

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from msb_extractor.export._styles import (
    DETAIL_FILL,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    STATUS_COLORS,
    TITLE_FONT,
    WRAP_ALIGN,
)
from msb_extractor.models import (
    ActualSet,
    DataSource,
    Exercise,
    ParseResult,
    PrescribedSet,
    SetStatus,
    TrainingDay,
)
from msb_extractor.normalize.exercise import apply_rename
from msb_extractor.normalize.units import Unit, format_load

_RAW_COLUMNS: list[tuple[str, str, int]] = [
    ("date", "Date", 12),
    ("day", "Day", 6),
    ("order", "Order", 7),
    ("exercise", "Exercise", 34),
    ("target", "Target (Prescribed)", 30),
    ("status", "Status", 12),
    ("set_number", "Set #", 7),
    ("reps", "Reps", 7),
    ("rpe", "RPE", 7),
    ("load", "Load", 14),
    ("percent_1rm", "%1RM", 8),
    ("e1rm", "e1RM", 14),
    ("comment", "Comment", 50),
    ("data_source", "Source", 14),
]

_INDEX_COLUMNS: list[tuple[str, int]] = [
    ("Exercise", 34),
    ("Total Sets", 12),
    ("Training Days", 14),
    ("First Seen", 14),
    ("Last Seen", 14),
]


def write_raw_log(
    ws: Worksheet,
    result: ParseResult,
    unit: Unit,
    rename_map: dict[str, str],
) -> None:
    for col_idx, (_key, label, width) in enumerate(_RAW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 2
    for day in result.days:
        for ex in day.exercises:
            display_name = apply_rename(ex.name, rename_map)
            for data in _expand_rows(day, ex, display_name):
                _write_row(ws, row, data, unit, day.data_source)
                row += 1

    last_row = max(row - 1, 1)
    last_col = get_column_letter(len(_RAW_COLUMNS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _expand_rows(
    day: TrainingDay,
    ex: Exercise,
    display_name: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    if ex.actuals:
        for idx, actual in enumerate(ex.actuals):
            pres = (
                ex.prescribed[idx]
                if idx < len(ex.prescribed)
                else (ex.prescribed[0] if ex.prescribed else None)
            )
            rows.append(_row_from_actual(day, ex, display_name, pres, actual))
    else:
        for pres in ex.prescribed:
            rows.append(_row_from_prescribed(day, ex, display_name, pres))

    if not rows:
        rows.append(
            {
                "date": day.date,
                "day": day.date.strftime("%a"),
                "order": ex.order,
                "exercise": display_name,
                "target": "",
                "status": "",
                "set_number": None,
                "reps": None,
                "rpe": None,
                "load": None,
                "percent_1rm": None,
                "e1rm": None,
                "comment": ex.notes or "",
                "data_source": day.data_source.value,
            }
        )
    return rows


def _row_from_actual(
    day: TrainingDay,
    ex: Exercise,
    display_name: str,
    pres: PrescribedSet | None,
    actual: ActualSet,
) -> dict[str, Any]:
    effective_status = (
        actual.status
        if actual.status != SetStatus.UNKNOWN
        else (pres.status if pres else SetStatus.UNKNOWN)
    )
    return {
        "date": day.date,
        "day": day.date.strftime("%a"),
        "order": ex.order,
        "exercise": display_name,
        "target": pres.target_text if pres else "",
        "status": effective_status.value,
        "status_enum": effective_status,
        "set_number": actual.set_number,
        "reps": actual.reps,
        "rpe": actual.rpe,
        "load": actual.load_kg,
        "load_display": actual.load_display,
        "percent_1rm": actual.percent_1rm,
        "e1rm": actual.e1rm_kg,
        "comment": actual.comment or "",
        "data_source": day.data_source.value,
    }


def _row_from_prescribed(
    day: TrainingDay,
    ex: Exercise,
    display_name: str,
    pres: PrescribedSet,
) -> dict[str, Any]:
    return {
        "date": day.date,
        "day": day.date.strftime("%a"),
        "order": ex.order,
        "exercise": display_name,
        "target": pres.target_text,
        "status": pres.status.value,
        "status_enum": pres.status,
        "set_number": None,
        "reps": pres.reps,
        "rpe": pres.rpe,
        "load": pres.load_kg,
        "load_display": pres.load_display,
        "percent_1rm": pres.percent_1rm,
        "e1rm": None,
        "comment": ex.notes or "",
        "data_source": day.data_source.value,
    }


def _write_row(
    ws: Worksheet,
    row: int,
    data: dict[str, Any],
    unit: Unit,
    source: DataSource,
) -> None:
    status_enum: SetStatus = data.get("status_enum", SetStatus.UNKNOWN)
    is_detail = source == DataSource.FULL_DETAIL

    for col_idx, (key, _label, _w) in enumerate(_RAW_COLUMNS, start=1):
        value: Any = _cell_value(key, data, unit)
        number_format = _cell_format(key)
        cell = ws.cell(row=row, column=col_idx, value=value)
        if number_format:
            cell.number_format = number_format
        if key == "comment":
            cell.alignment = WRAP_ALIGN
        if is_detail:
            cell.fill = DETAIL_FILL
        if key == "status" and status_enum in STATUS_COLORS:
            cell.font = Font(color=STATUS_COLORS[status_enum], bold=True)


def _cell_value(key: str, data: dict[str, Any], unit: Unit) -> Any:
    if key == "date":
        return data["date"]
    if key == "load":
        load_kg = data.get("load")
        display = data.get("load_display")
        if load_kg is not None:
            return format_load(load_kg, unit, decimals=1)
        return display or ""
    if key == "e1rm":
        return format_load(data["e1rm"], unit, decimals=1) if data["e1rm"] else ""
    if key == "status":
        return data["status"] or ""
    if key in ("day", "order", "exercise", "target", "comment", "data_source"):
        return data[key]
    return data.get(key)


def _cell_format(key: str) -> str | None:
    if key == "date":
        return "yyyy-mm-dd"
    if key == "rpe":
        return "0.0"
    if key == "percent_1rm":
        return "0%"
    return None


def write_exercise_index(
    ws: Worksheet,
    result: ParseResult,
    rename_map: dict[str, str],
) -> None:
    stats: dict[str, dict[str, Any]] = {}
    for day in result.days:
        for ex in day.exercises:
            name = apply_rename(ex.name, rename_map)
            entry = stats.setdefault(
                name,
                {"sets": 0, "days": set(), "first": day.date, "last": day.date},
            )
            entry["sets"] += max(len(ex.actuals), len(ex.prescribed))
            entry["days"].add(day.date)
            entry["first"] = min(entry["first"], day.date)
            entry["last"] = max(entry["last"], day.date)

    for col_idx, (label, width) in enumerate(_INDEX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ordered = sorted(stats.items(), key=lambda kv: (-kv[1]["sets"], kv[0]))
    for i, (name, entry) in enumerate(ordered, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=entry["sets"])
        ws.cell(row=i, column=3, value=len(entry["days"]))
        first_cell = ws.cell(row=i, column=4, value=entry["first"])
        first_cell.number_format = "yyyy-mm-dd"
        last_cell = ws.cell(row=i, column=5, value=entry["last"])
        last_cell.number_format = "yyyy-mm-dd"

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_INDEX_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(ordered) + 1, 1)}"


def write_summary(
    ws: Worksheet,
    result: ParseResult,
    unit: Unit,
    rename_map: dict[str, str],
) -> None:
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60

    title_cell = ws.cell(row=1, column=1, value="MSB Extractor - Training Log Summary")
    title_cell.font = TITLE_FONT

    date_range = result.date_range
    first = date_range[0].isoformat() if date_range else "n/a"
    last = date_range[1].isoformat() if date_range else "n/a"

    source_counts = Counter(d.data_source for d in result.days)

    rows: list[tuple[str, Any]] = [
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("Captured at", result.captured_at.isoformat() if result.captured_at else "n/a"),
        ("Source", result.source),
        ("Output unit", unit),
        ("", ""),
        ("First training day", first),
        ("Last training day", last),
        ("Total training days", len(result.days)),
        ("Days with full detail", source_counts.get(DataSource.FULL_DETAIL, 0)),
        ("Days with calendar-only data", source_counts.get(DataSource.CALENDAR, 0)),
        ("Total sets", result.total_sets),
        ("Unique exercises", len(result.exercise_names)),
        ("", ""),
        ("Legend: row shaded green", "Full-detail day (actuals, RPE, comments)"),
        ("Legend: white row", "Calendar-only day (prescription, partial actuals)"),
    ]

    for row_idx, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        if label:
            label_cell.font = LABEL_FONT
        ws.cell(row=row_idx, column=2, value=value)

    if rename_map:
        rename_row = len(rows) + 4
        ws.cell(row=rename_row, column=1, value="Applied rename map").font = LABEL_FONT
        ws.cell(row=rename_row, column=2, value=f"{len(rename_map)} entries")

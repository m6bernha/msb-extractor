"""Main xlsx export orchestrator.

Sheet order:
    Raw Log             flat, one row per set; the query surface
    Week ...            one sheet per ISO week, wide format (hero sheet)
    Exercise Progress   top-set-per-day per exercise, for progress charts
    Exercise Index      per-exercise totals and date range
    Summary             coverage report and legend

Only Raw Log + Summary are guaranteed; the rest can be opted out via flags.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from msb_extractor.export._charts import write_progress_charts
from msb_extractor.export._flat import (
    write_exercise_index,
    write_raw_log,
    write_summary,
)
from msb_extractor.export._progress import ProgressRange, write_exercise_progress
from msb_extractor.export._weekly import write_weekly_view
from msb_extractor.models import ParseResult
from msb_extractor.normalize.exercise import load_rename_map
from msb_extractor.normalize.units import Unit


def write_xlsx(
    result: ParseResult,
    output_path: str | Path,
    unit: Unit = "kg",
    rename_map_path: str | Path | None = None,
    include_weekly: bool = True,
    include_progress: bool = True,
    include_charts: bool = True,
) -> Path:
    """Write a parsed training log to an xlsx file and return the path."""
    rename_map = load_rename_map(rename_map_path)
    wb = Workbook()

    ws_raw = wb.active
    if ws_raw is None:
        ws_raw = wb.create_sheet("Raw Log")
    else:
        ws_raw.title = "Raw Log"
    write_raw_log(ws_raw, result, unit, rename_map)

    if include_weekly:
        write_weekly_view(wb, result, unit, rename_map)

    progress_ranges: dict[str, ProgressRange] = {}
    if include_progress:
        ws_progress = wb.create_sheet("Exercise Progress")
        progress_ranges = write_exercise_progress(ws_progress, result, unit, rename_map)

    if include_charts and include_progress and progress_ranges:
        write_progress_charts(wb, progress_ranges, unit)

    ws_index = wb.create_sheet("Exercise Index")
    write_exercise_index(ws_index, result, rename_map)

    ws_summary = wb.create_sheet("Summary")
    write_summary(ws_summary, result, unit, rename_map)

    path = Path(output_path)
    wb.save(path)
    return path

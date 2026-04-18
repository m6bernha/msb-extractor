"""Shared cell styles used by all xlsx sheet writers."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from msb_extractor.models import SetStatus

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
SUBHEADER_FONT = Font(bold=True, color="F9FAFB", size=10)

DAY_BANNER_FILL = PatternFill("solid", fgColor="E5E7EB")
DAY_BANNER_FONT = Font(bold=True, size=11, color="111827")

DETAIL_FILL = PatternFill("solid", fgColor="DCFCE7")
CALENDAR_FILL = PatternFill("solid", fgColor="F3F4F6")

TITLE_FONT = Font(bold=True, size=14, color="1F2937")
LABEL_FONT = Font(bold=True, color="1F2937")

_THIN = Side(style="thin", color="D1D5DB")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

STATUS_COLORS: dict[SetStatus, str] = {
    SetStatus.COMPLETED: "166534",
    SetStatus.PARTIAL: "92400E",
    SetStatus.MISSED: "991B1B",
    SetStatus.PRESCRIBED: "1D4ED8",
    SetStatus.UNKNOWN: "6B7280",
}

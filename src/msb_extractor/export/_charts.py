"""Line charts for estimated 1RM over time, one chart per eligible exercise.

Charts are built on top of the Exercise Progress sheet's numeric e1RM column
so the chart data stays queryable. Charts land on a dedicated ``e1RM Charts``
sheet in a 2-wide grid.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList

from msb_extractor.export._progress import E1RM_COLUMN, ProgressRange
from msb_extractor.normalize.units import Unit

_CHARTS_PER_ROW = 2
_CHART_WIDTH = 14
_CHART_HEIGHT = 8
_CHART_COL_SPAN = 8
_CHART_ROW_SPAN = 17
_MIN_E1RM_POINTS = 5


def write_progress_charts(
    wb: Workbook,
    ranges: dict[str, ProgressRange],
    unit: Unit,
    progress_sheet_name: str = "Exercise Progress",
) -> None:
    """Emit a line chart of e1RM vs date for every exercise with enough data."""
    eligible = [
        (name, rng) for name, rng in sorted(ranges.items()) if rng.e1rm_points >= _MIN_E1RM_POINTS
    ]
    if not eligible:
        return

    progress_ws = wb[progress_sheet_name]
    charts_ws = wb.create_sheet("e1RM Charts")
    charts_ws.column_dimensions["A"].width = 2

    title = charts_ws.cell(row=1, column=2, value="Estimated 1RM over time")
    title.font = _title_font()
    charts_ws.cell(
        row=2,
        column=2,
        value=(
            f"Unit: {unit}. Each chart uses the best e1RM MSB logged per training day, "
            f"drawn from the Exercise Progress sheet. Only exercises with "
            f"{_MIN_E1RM_POINTS}+ datapoints are shown."
        ),
    )

    for idx, (exercise_name, rng) in enumerate(eligible):
        chart = LineChart()
        chart.title = exercise_name
        chart.width = _CHART_WIDTH
        chart.height = _CHART_HEIGHT
        chart.legend = None
        chart.y_axis.title = f"e1RM ({unit})"
        chart.x_axis.title = "Date"
        chart.x_axis.number_format = "yyyy-mm-dd"

        data_ref = Reference(
            progress_ws,
            min_col=E1RM_COLUMN,
            max_col=E1RM_COLUMN,
            min_row=rng.first_row,
            max_row=rng.last_row,
        )
        chart.add_data(data_ref, titles_from_data=False)

        date_ref = Reference(
            progress_ws,
            min_col=2,
            max_col=2,
            min_row=rng.first_row,
            max_row=rng.last_row,
        )
        chart.set_categories(date_ref)

        series = chart.series[0]
        series.smooth = False
        chart.dataLabels = DataLabelList(showVal=False)

        row = 4 + (idx // _CHARTS_PER_ROW) * _CHART_ROW_SPAN
        col = 2 + (idx % _CHARTS_PER_ROW) * _CHART_COL_SPAN
        anchor_cell = charts_ws.cell(row=row, column=col).coordinate
        charts_ws.add_chart(chart, anchor_cell)


def _title_font() -> object:
    from msb_extractor.export._styles import TITLE_FONT

    return TITLE_FONT

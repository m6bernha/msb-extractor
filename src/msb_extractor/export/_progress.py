"""Exercise-Progress sheet writer.

One row per (exercise, training day): the top load, the best e1RM of the
day, total sets logged, and what reps / RPE you hit on the top set. The
e1RM is stored as a numeric value with a unit-aware display format so the
charts writer can reference it directly.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from msb_extractor.export._styles import HEADER_ALIGN, HEADER_FILL, HEADER_FONT
from msb_extractor.models import ActualSet, ParseResult
from msb_extractor.normalize.exercise import apply_rename
from msb_extractor.normalize.units import Unit, convert_kg, format_load

_COLUMNS: list[tuple[str, int]] = [
    ("Exercise", 34),
    ("Date", 12),
    ("Day", 6),
    ("Top Load", 14),
    ("Top Reps", 9),
    ("Top RPE", 9),
    ("Best e1RM", 14),
    ("Sets", 7),
]

E1RM_COLUMN = 7


@dataclass(frozen=True)
class ProgressRange:
    """Row range a single exercise occupies in the Exercise Progress sheet."""

    first_row: int
    last_row: int
    e1rm_points: int

    @property
    def row_count(self) -> int:
        return self.last_row - self.first_row + 1


def write_exercise_progress(
    ws: Worksheet,
    result: ParseResult,
    unit: Unit,
    rename_map: dict[str, str],
) -> dict[str, ProgressRange]:
    """Write the sheet and return per-exercise row ranges.

    The returned dict maps each exercise name to the row range it occupies,
    along with how many of those rows carry a numeric e1RM value. Callers
    (the charts writer) use this to position chart series without scanning
    the sheet again.
    """
    for col_idx, (label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    rows: list[dict[str, Any]] = []
    for day in result.days:
        for ex in day.exercises:
            if not ex.actuals:
                continue
            top = _pick_top_set(ex.actuals)
            best_e1rm = max(
                (a.e1rm_kg for a in ex.actuals if a.e1rm_kg is not None),
                default=None,
            )
            rows.append(
                {
                    "exercise": apply_rename(ex.name, rename_map),
                    "date": day.date,
                    "day": day.date.strftime("%a"),
                    "top_load": top.load_kg,
                    "top_reps": top.reps,
                    "top_rpe": top.rpe,
                    "best_e1rm": best_e1rm,
                    "sets": len(ex.actuals),
                }
            )

    rows.sort(key=lambda r: (r["exercise"], r["date"]))

    e1rm_format = f'0.0 "{unit}"'
    ranges: dict[str, ProgressRange] = {}
    current_exercise: str | None = None
    current_start: int = 2
    current_e1rm_count: int = 0

    for i, r in enumerate(rows, start=2):
        if r["exercise"] != current_exercise:
            if current_exercise is not None:
                ranges[current_exercise] = ProgressRange(
                    first_row=current_start,
                    last_row=i - 1,
                    e1rm_points=current_e1rm_count,
                )
            current_exercise = r["exercise"]
            current_start = i
            current_e1rm_count = 0

        ws.cell(row=i, column=1, value=r["exercise"])
        date_cell = ws.cell(row=i, column=2, value=r["date"])
        date_cell.number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=3, value=r["day"])
        ws.cell(
            row=i,
            column=4,
            value=format_load(r["top_load"], unit, decimals=1) if r["top_load"] else "",
        )
        ws.cell(row=i, column=5, value=r["top_reps"])
        rpe_cell = ws.cell(row=i, column=6, value=r["top_rpe"])
        if r["top_rpe"] is not None:
            rpe_cell.number_format = "0.0"

        if r["best_e1rm"] is not None:
            e1rm_value = convert_kg(r["best_e1rm"], unit)
            e1rm_cell = ws.cell(row=i, column=E1RM_COLUMN, value=e1rm_value)
            e1rm_cell.number_format = e1rm_format
            current_e1rm_count += 1
        else:
            ws.cell(row=i, column=E1RM_COLUMN, value=None)

        ws.cell(row=i, column=8, value=r["sets"])

    if current_exercise is not None:
        ranges[current_exercise] = ProgressRange(
            first_row=current_start,
            last_row=len(rows) + 1,
            e1rm_points=current_e1rm_count,
        )

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"
    return ranges


def _pick_top_set(actuals: list[ActualSet]) -> ActualSet:
    """The set with the highest load; ties broken by reps then set order."""
    return max(actuals, key=lambda a: (a.load_kg or 0, a.reps or 0, -a.set_number))
